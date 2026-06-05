from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session, joinedload
from typing import List
from fastapi.responses import StreamingResponse
import pandas as pd
import io
from datetime import datetime

from database import get_db
from models import Dimension, Subdimension, EvaluacionRespuesta, Usuario
from schemas import (
    DimensionCreate, DimensionUpdate, DimensionResponse,
    SubdimensionCreate, SubdimensionUpdate, SubdimensionResponse,
    ReorderRequest
)
from auth import get_current_active_user, require_admin

router = APIRouter(prefix="/dimensiones", tags=["Dimensiones y Subdimensiones"])


@router.get("/", response_model=List[DimensionResponse])
def list_dimensiones(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user)
):
    dimensiones = db.query(Dimension).options(
        joinedload(Dimension.subdimensiones)
    ).order_by(Dimension.orden).all()
    
    result = []
    for dim in dimensiones:
        result.append({
            "id": dim.id,
            "nombre": dim.nombre,
            "descripcion": dim.descripcion,
            "orden": dim.orden,
            "subdimensiones": sorted([
                {
                    "id": sub.id,
                    "dimension_id": sub.dimension_id,
                    "nombre": sub.nombre,
                    "descripcion": sub.descripcion,
                    "orden": sub.orden
                }
                for sub in dim.subdimensiones
            ], key=lambda x: x["orden"])
        })
    return result


@router.post("/", response_model=DimensionResponse)
def create_dimension(
    dimension: DimensionCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    max_orden = db.query(Dimension).order_by(Dimension.orden.desc()).first()
    nuevo_orden = (max_orden.orden + 1) if max_orden else 1
    
    db_dimension = Dimension(
        nombre=dimension.nombre,
        descripcion=dimension.descripcion,
        orden=nuevo_orden
    )
    db.add(db_dimension)
    db.commit()
    db.refresh(db_dimension)
    
    return {
        "id": db_dimension.id,
        "nombre": db_dimension.nombre,
        "descripcion": db_dimension.descripcion,
        "orden": db_dimension.orden,
        "subdimensiones": []
    }


@router.put("/{dimension_id}", response_model=DimensionResponse)
def update_dimension(
    dimension_id: int,
    dimension_update: DimensionUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    db_dimension = db.query(Dimension).filter(Dimension.id == dimension_id).first()
    if not db_dimension:
        raise HTTPException(status_code=404, detail="Dimensión no encontrada")
    
    if dimension_update.nombre is not None:
        db_dimension.nombre = dimension_update.nombre
    if dimension_update.descripcion is not None:
        db_dimension.descripcion = dimension_update.descripcion
    
    db.commit()
    db.refresh(db_dimension)
    
    return {
        "id": db_dimension.id,
        "nombre": db_dimension.nombre,
        "descripcion": db_dimension.descripcion,
        "orden": db_dimension.orden,
        "subdimensiones": sorted([
            {
                "id": sub.id,
                "dimension_id": sub.dimension_id,
                "nombre": sub.nombre,
                "descripcion": sub.descripcion,
                "orden": sub.orden
            }
            for sub in db_dimension.subdimensiones
        ], key=lambda x: x["orden"])
    }


@router.delete("/{dimension_id}")
def delete_dimension(
    dimension_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    db_dimension = db.query(Dimension).filter(Dimension.id == dimension_id).first()
    if not db_dimension:
        raise HTTPException(status_code=404, detail="Dimensión no encontrada")
    
    subdims = db.query(Subdimension).filter(Subdimension.dimension_id == dimension_id).all()
    for sub in subdims:
        respuestas = db.query(EvaluacionRespuesta).filter(
            EvaluacionRespuesta.subdimension_id == sub.id
        ).count()
        if respuestas > 0:
            raise HTTPException(
                status_code=400,
                detail=f"No se puede eliminar la dimensión. El indicador '{sub.nombre}' tiene evaluaciones asociadas."
            )
    
    db.query(Subdimension).filter(Subdimension.dimension_id == dimension_id).delete()
    db.delete(db_dimension)
    db.commit()
    
    return {"message": "Dimensión eliminada correctamente"}


@router.put("/reorder")
def reorder_dimensiones(
    reorder_data: ReorderRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    for idx, dim_id in enumerate(reorder_data.ids):
        dim = db.query(Dimension).filter(Dimension.id == dim_id).first()
        if dim:
            dim.orden = idx + 1
    
    db.commit()
    return {"message": "Orden actualizado correctamente"}


# SUBDIMENSIONES (Indicadores)

@router.get("/subdimensiones", response_model=List[SubdimensionResponse])
def list_subdimensiones(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user)
):
    return db.query(Subdimension).order_by(Subdimension.orden).all()


@router.post("/{dimension_id}/subdimensiones/", response_model=SubdimensionResponse)
def create_subdimension(
    dimension_id: int,
    subdimension: SubdimensionCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    dimension = db.query(Dimension).filter(Dimension.id == dimension_id).first()
    if not dimension:
        raise HTTPException(status_code=404, detail="Dimensión no encontrada")
    
    max_orden = db.query(Subdimension).filter(
        Subdimension.dimension_id == dimension_id
    ).order_by(Subdimension.orden.desc()).first()
    nuevo_orden = (max_orden.orden + 1) if max_orden else 1
    
    db_sub = Subdimension(
        dimension_id=dimension_id,
        nombre=subdimension.nombre,
        descripcion=subdimension.descripcion,
        orden=nuevo_orden
    )
    db.add(db_sub)
    db.commit()
    db.refresh(db_sub)
    
    return {
        "id": db_sub.id,
        "dimension_id": db_sub.dimension_id,
        "nombre": db_sub.nombre,
        "descripcion": db_sub.descripcion,
        "orden": db_sub.orden
    }


@router.put("/subdimensiones/{subdimension_id}", response_model=SubdimensionResponse)
def update_subdimension(
    subdimension_id: int,
    subdimension_update: SubdimensionUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    db_sub = db.query(Subdimension).filter(Subdimension.id == subdimension_id).first()
    if not db_sub:
        raise HTTPException(status_code=404, detail="Indicador no encontrado")
    
    if subdimension_update.nombre is not None:
        db_sub.nombre = subdimension_update.nombre
    if subdimension_update.descripcion is not None:
        db_sub.descripcion = subdimension_update.descripcion
    
    db.commit()
    db.refresh(db_sub)
    
    return {
        "id": db_sub.id,
        "dimension_id": db_sub.dimension_id,
        "nombre": db_sub.nombre,
        "descripcion": db_sub.descripcion,
        "orden": db_sub.orden
    }


@router.delete("/subdimensiones/{subdimension_id}")
def delete_subdimension(
    subdimension_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    db_sub = db.query(Subdimension).filter(Subdimension.id == subdimension_id).first()
    if not db_sub:
        raise HTTPException(status_code=404, detail="Indicador no encontrado")
    
    respuestas = db.query(EvaluacionRespuesta).filter(
        EvaluacionRespuesta.subdimension_id == subdimension_id
    ).count()
    
    if respuestas > 0:
        raise HTTPException(
            status_code=400,
            detail=f"No se puede eliminar el indicador '{db_sub.nombre}' porque tiene evaluaciones asociadas."
        )
    
    db.delete(db_sub)
    db.commit()
    
    return {"message": "Indicador eliminado correctamente"}


@router.put("/subdimensiones/reorder")
def reorder_subdimensiones(
    reorder_data: ReorderRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    for idx, sub_id in enumerate(reorder_data.ids):
        sub = db.query(Subdimension).filter(Subdimension.id == sub_id).first()
        if sub:
            sub.orden = idx + 1
    
    db.commit()
    return {"message": "Orden de indicadores actualizado correctamente"}


# EXPORTACIÓN

@router.get("/export/excel")
def exportar_plantilla_excel(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user)
):
    dimensiones = db.query(Dimension).options(
        joinedload(Dimension.subdimensiones)
    ).order_by(Dimension.orden).all()
    
    data = []
    for dim in dimensiones:
        subdims_ordenadas = sorted(dim.subdimensiones, key=lambda x: x.orden)
        if subdims_ordenadas:
            for sub in subdims_ordenadas:
                data.append({
                    "Dimensión": dim.nombre,
                    "Indicador": sub.nombre,
                    "Descripción": sub.descripcion or "",
                    "Orden": sub.orden
                })
        else:
            data.append({
                "Dimensión": dim.nombre,
                "Indicador": "(Sin indicadores)",
                "Descripción": "",
                "Orden": ""
            })
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Plantilla")
        
        worksheet = writer.sheets["Plantilla"]
        for idx in range(len(df.columns)):
            col_letter = chr(65 + idx)
            max_length = max(
                df.iloc[:, idx].astype(str).map(len).max() if len(df) > 0 else 0,
                len(df.columns[idx])
            ) + 2
            worksheet.column_dimensions[col_letter].width = min(max_length, 60)
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="002B5E", end_color="002B5E", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    output.seek(0)
    filename = f"plantilla_liderazgo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
