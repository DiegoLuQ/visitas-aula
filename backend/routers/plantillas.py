import re
import io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, func
from typing import List, Optional
import pandas as pd

from database import get_db
from models import Plantilla, Dimension, Subdimension, Usuario, Evaluacion
from schemas import PlantillaCreate, PlantillaResponse, DimensionResponse, PlantillaUpdate, PlantillaDuplicate
from auth import get_current_active_user, require_admin, require_admin_or_director

router = APIRouter(prefix="/eval_plantillas", tags=["Plantillas de Evaluación"])


def _is_admin(user: Usuario) -> bool:
    return user.rol_id == 1


def _colegios_usuario(user: Usuario) -> list[int]:
    """Lista de IDs de colegio asignados al usuario (Usuario.colegio_id puede ser '1' o '1,2')."""
    if not user.colegio_id:
        return []
    return [int(x.strip()) for x in str(user.colegio_id).split(",") if x.strip().isdigit()]


def _es_director(user: Usuario) -> bool:
    return bool(user.rol and (user.rol.nombre or "").lower() == "director")


def _verificar_acceso_plantilla(user: Usuario, plantilla: Plantilla) -> None:
    """El director solo puede acceder a plantillas de su(s) colegio(s)."""
    if _is_admin(user) or not _es_director(user):
        return
    colegios = _colegios_usuario(user)
    if plantilla.colegio_id is None or plantilla.colegio_id not in colegios:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta plantilla")


def _slug_unico(db: Session, base: str) -> str:
    """Genera un slug único (<=20 chars) a partir de un texto base."""
    base = re.sub(r"[^a-z0-9]+", "-", (base or "plantilla").lower()).strip("-")[:15] or "plantilla"
    slug = base
    i = 1
    while db.query(Plantilla).filter(Plantilla.slug == slug).first():
        suffix = f"-{i}"
        slug = base[: 20 - len(suffix)] + suffix
        i += 1
    return slug[:20]


@router.get("/", response_model=List[PlantillaResponse])
def list_plantillas(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user)
):
    query = db.query(Plantilla).filter(Plantilla.activa == True)

    if not _is_admin(current_user):
        colegios = _colegios_usuario(current_user)
        role_name = (current_user.rol.nombre or "").lower() if current_user.rol else ""

        if role_name == "director":
            # El director SOLO ve las plantillas de su(s) colegio(s), ninguna otra
            # (ni globales/LIDERAZGO ni de otros colegios).
            if colegios:
                query = query.filter(Plantilla.colegio_id.in_(colegios))
            else:
                # Sin colegio asignado => no ve plantillas
                query = query.filter(Plantilla.id == -1)
        elif colegios:
            # Otros roles con colegio: ven las globales (LIDERAZGO) + las VISITA de su(s) colegio(s).
            query = query.filter(
                or_(
                    func.upper(Plantilla.tipo) == "LIDERAZGO",
                    Plantilla.colegio_id.in_(colegios),
                )
            )
    return query.all()

@router.get("/{plantilla_id}", response_model=PlantillaResponse)
def get_plantilla(
    plantilla_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user)
):
    plantilla = db.query(Plantilla).filter(Plantilla.id == plantilla_id).first()
    if not plantilla:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    _verificar_acceso_plantilla(current_user, plantilla)
    return plantilla

@router.get("/{plantilla_id}/dimensiones", response_model=List[DimensionResponse])
def get_plantilla_dimensiones(
    plantilla_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user)
):
    plantilla = db.query(Plantilla).filter(Plantilla.id == plantilla_id).first()
    if not plantilla:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    _verificar_acceso_plantilla(current_user, plantilla)

    dimensiones = db.query(Dimension).filter(
        Dimension.plantilla_id == plantilla_id
    ).options(
        joinedload(Dimension.subdimensiones)
    ).order_by(Dimension.orden).all()
    
    result = []
    for dim in dimensiones:
        result.append({
            "id": dim.id,
            "plantilla_id": dim.plantilla_id,
            "nombre": dim.nombre,
            "descripcion": dim.descripcion,
            "orden": dim.orden,
            "subdimensiones": sorted([
                {
                    "id": sub.id,
                    "dimension_id": sub.dimension_id,
                    "nombre": sub.nombre,
                    "descripcion": sub.descripcion,
                    "orden": sub.orden
                }
                for sub in dim.subdimensiones
            ], key=lambda x: x["orden"])
        })
    return result

@router.post("/", response_model=PlantillaResponse)
def create_plantilla(
    plantilla: PlantillaCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    db_plantilla = Plantilla(**plantilla.dict())
    db.add(db_plantilla)
    db.commit()
    db.refresh(db_plantilla)
    return db_plantilla


@router.put("/{plantilla_id}", response_model=PlantillaResponse)
def update_plantilla(
    plantilla_id: int,
    plantilla_update: PlantillaUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    db_plantilla = db.query(Plantilla).filter(Plantilla.id == plantilla_id).first()
    if not db_plantilla:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    
    if plantilla_update.nombre is not None:
        db_plantilla.nombre = plantilla_update.nombre
    if plantilla_update.nombre_largo is not None:
        db_plantilla.nombre_largo = plantilla_update.nombre_largo
    if plantilla_update.slug is not None:
        db_plantilla.slug = plantilla_update.slug
    if plantilla_update.tipo is not None:
        db_plantilla.tipo = plantilla_update.tipo
    if plantilla_update.formato is not None:
        db_plantilla.formato = plantilla_update.formato
    if plantilla_update.colegio_id is not None:
        # colegio_id = 0 lo interpretamos como "sin colegio" (global / sin asignar)
        db_plantilla.colegio_id = plantilla_update.colegio_id or None
    if plantilla_update.config_puntuacion is not None:
        db_plantilla.config_puntuacion = plantilla_update.config_puntuacion
    if plantilla_update.activa is not None:
        db_plantilla.activa = plantilla_update.activa

    db.commit()
    db.refresh(db_plantilla)
    return db_plantilla


@router.post("/{plantilla_id}/duplicar", response_model=PlantillaResponse)
def duplicar_plantilla(
    plantilla_id: int,
    payload: PlantillaDuplicate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin_or_director)
):
    original = db.query(Plantilla).filter(Plantilla.id == plantilla_id).first()
    if not original:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")

    # Determinar el colegio destino de la copia
    if _is_admin(current_user):
        # El admin debe indicar a qué colegio pertenece la nueva plantilla
        if not payload.colegio_id:
            raise HTTPException(status_code=400, detail="Debe seleccionar el colegio destino de la plantilla")
        target_colegio = payload.colegio_id
    else:
        # Director: se asigna a su colegio. Si tiene varios, debe elegir uno de los suyos.
        colegios = _colegios_usuario(current_user)
        if not colegios:
            raise HTTPException(status_code=400, detail="No tiene un colegio asignado para crear la plantilla")
        if len(colegios) == 1:
            target_colegio = colegios[0]
        else:
            if not payload.colegio_id or payload.colegio_id not in colegios:
                raise HTTPException(status_code=400, detail="Debe seleccionar uno de sus colegios asignados")
            target_colegio = payload.colegio_id

    # Crear la nueva plantilla (cabecera)
    nueva = Plantilla(
        nombre=f"{original.nombre} (copia)",
        nombre_largo=original.nombre_largo,
        slug=_slug_unico(db, original.slug or original.nombre),
        tipo=original.tipo,
        formato=original.formato,
        colegio_id=target_colegio,
        config_puntuacion=original.config_puntuacion,
        activa=True,
    )
    db.add(nueva)
    db.flush()  # obtiene nueva.id sin cerrar la transacción

    # Clonar dimensiones y sus subdimensiones
    dimensiones = db.query(Dimension).filter(Dimension.plantilla_id == original.id).all()
    for dim in dimensiones:
        nueva_dim = Dimension(
            plantilla_id=nueva.id,
            nombre=dim.nombre,
            descripcion=dim.descripcion,
            orden=dim.orden,
        )
        db.add(nueva_dim)
        db.flush()
        subs = db.query(Subdimension).filter(Subdimension.dimension_id == dim.id).all()
        for sub in subs:
            db.add(Subdimension(
                dimension_id=nueva_dim.id,
                nombre=sub.nombre,
                descripcion=sub.descripcion,
                orden=sub.orden,
            ))

    db.commit()
    db.refresh(nueva)
    return nueva


@router.delete("/{plantilla_id}")
def delete_plantilla(
    plantilla_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin_or_director)
):
    db_plantilla = db.query(Plantilla).filter(Plantilla.id == plantilla_id).first()
    if not db_plantilla:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")

    # Un director solo puede eliminar plantillas de su(s) colegio(s)
    if not _is_admin(current_user):
        colegios = _colegios_usuario(current_user)
        if db_plantilla.colegio_id is None or db_plantilla.colegio_id not in colegios:
            raise HTTPException(status_code=403, detail="Solo puedes eliminar plantillas de tu colegio")

    # No se puede eliminar si está en uso (tiene evaluaciones/visitas asociadas)
    en_uso = db.query(Evaluacion).filter(Evaluacion.plantilla_id == plantilla_id).count()
    if en_uso > 0:
        raise HTTPException(
            status_code=409,
            detail=f"No se puede eliminar: la plantilla está en uso ({en_uso} registro(s) asociado(s)). Desactívala o elimina primero esos registros."
        )

    # Eliminación definitiva: indicadores -> dimensiones -> plantilla
    dimensiones = db.query(Dimension).filter(Dimension.plantilla_id == plantilla_id).all()
    for dim in dimensiones:
        db.query(Subdimension).filter(Subdimension.dimension_id == dim.id).delete(synchronize_session=False)
        db.delete(dim)
    db.delete(db_plantilla)
    db.commit()
    return {"message": "Plantilla eliminada correctamente"}


@router.get("/{plantilla_id}/export/excel")
def export_plantilla_estructura(
    plantilla_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user)
):
    """Exporta la estructura (dimensiones + indicadores) de UNA plantilla como Excel."""
    plantilla = db.query(Plantilla).filter(Plantilla.id == plantilla_id).first()
    if not plantilla:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")

    dimensiones = db.query(Dimension).filter(
        Dimension.plantilla_id == plantilla_id
    ).options(joinedload(Dimension.subdimensiones)).order_by(Dimension.orden).all()

    data = []
    for dim in dimensiones:
        subs = sorted(dim.subdimensiones, key=lambda x: x.orden)
        if subs:
            for sub in subs:
                data.append({
                    "Dimensión": dim.nombre,
                    "Indicador": sub.nombre,
                    "Descripción": sub.descripcion or "",
                    "Orden": sub.orden,
                })
        else:
            data.append({"Dimensión": dim.nombre, "Indicador": "(Sin indicadores)", "Descripción": "", "Orden": ""})

    if not data:
        # Plantilla sin estructura: entregar al menos los encabezados de ejemplo
        data = [{"Dimensión": "", "Indicador": "", "Descripción": "", "Orden": ""}]

    df = pd.DataFrame(data, columns=["Dimensión", "Indicador", "Descripción", "Orden"])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Estructura")
        ws = writer.sheets["Estructura"]
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="002B5E", end_color="002B5E", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 8

    output.seek(0)
    filename = f"estructura_{(plantilla.slug or 'plantilla')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/import/excel", response_model=PlantillaResponse)
def import_plantilla_excel(
    nombre: str = Form(...),
    tipo: str = Form(...),
    slug: str = Form(""),
    formato: Optional[str] = Form(None),
    nombre_largo: Optional[str] = Form(None),
    colegio_id: Optional[int] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_admin_or_director)
):
    """Crea una NUEVA plantilla a partir de la estructura (dimensiones + indicadores) de un Excel."""
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Formato no válido. Use un archivo Excel (.xlsx).")

    # Resolver colegio destino según el rol (mismo criterio que duplicar)
    if _is_admin(current_user):
        target_colegio = colegio_id or None  # admin puede dejar la plantilla sin colegio (global)
    else:
        colegios = _colegios_usuario(current_user)
        if not colegios:
            raise HTTPException(status_code=400, detail="No tiene un colegio asignado")
        if colegio_id and colegio_id in colegios:
            target_colegio = colegio_id
        elif len(colegios) == 1:
            target_colegio = colegios[0]
        else:
            raise HTTPException(status_code=400, detail="Debe seleccionar uno de sus colegios asignados")

    # Slug: si viene vacío se genera; si viene, debe ser único
    slug = (slug or "").strip()
    if not slug:
        slug = _slug_unico(db, nombre)
    elif db.query(Plantilla).filter(Plantilla.slug == slug).first():
        raise HTTPException(status_code=400, detail=f"El slug '{slug}' ya existe. Usa otro.")

    # Leer el Excel
    try:
        contents = file.file.read()
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")

    # Mapear columnas de forma tolerante (acentos / mayúsculas)
    cols = {str(c).strip().lower(): c for c in df.columns}
    col_dim = cols.get('dimensión') or cols.get('dimension')
    col_ind = cols.get('indicador')
    col_desc = cols.get('descripción') or cols.get('descripcion')
    if not col_dim or not col_ind:
        raise HTTPException(status_code=400, detail="El Excel debe tener al menos las columnas 'Dimensión' e 'Indicador'.")

    # Crear cabecera de la plantilla
    nueva = Plantilla(
        nombre=nombre.strip(),
        nombre_largo=(nombre_largo or "").strip() or None,
        slug=slug,
        tipo=tipo.strip(),
        formato=(formato or "").strip() or None,
        colegio_id=target_colegio,
        activa=True,
    )
    db.add(nueva)
    db.flush()

    # Construir dimensiones (por primera aparición) e indicadores (en orden de fila)
    dim_cache = {}
    dim_orden = 0
    ind_orden = {}
    for _, row in df.iterrows():
        dim_nombre = str(row[col_dim]).strip() if pd.notna(row[col_dim]) else ''
        if not dim_nombre:
            continue
        if dim_nombre not in dim_cache:
            dim_orden += 1
            d = Dimension(plantilla_id=nueva.id, nombre=dim_nombre, orden=dim_orden)
            db.add(d)
            db.flush()
            dim_cache[dim_nombre] = d
            ind_orden[dim_nombre] = 0

        ind_nombre = str(row[col_ind]).strip() if pd.notna(row[col_ind]) else ''
        if not ind_nombre or ind_nombre == '(Sin indicadores)':
            continue
        desc = str(row[col_desc]).strip() if (col_desc and pd.notna(row[col_desc])) else ''
        ind_orden[dim_nombre] += 1
        db.add(Subdimension(
            dimension_id=dim_cache[dim_nombre].id,
            nombre=ind_nombre,
            descripcion=desc or None,
            orden=ind_orden[dim_nombre],
        ))

    if not dim_cache:
        db.rollback()
        raise HTTPException(status_code=400, detail="El Excel no contiene dimensiones válidas.")

    db.commit()
    db.refresh(nueva)
    return nueva
